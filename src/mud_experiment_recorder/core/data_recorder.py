from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl


class ExcelRecorder:
    """Generates multi-sheet .xlsx from session CSV on recording stop.

    CSV writing is handled by URFTAdapter during recording (fast path).
    This class does the post-hoc CSV -> .xlsx conversion.
    """

    def __init__(self) -> None:
        self.session_dir: Optional[Path] = None

    def finalize(
        self,
        session_dir: Path,
        csv_path: Path,
        sample_count: int,
        weight_history: List[tuple] = None,
    ) -> Optional[Path]:
        """Convert CSV to .xlsx. Returns xlsx_path or None on failure."""
        self.session_dir = session_dir
        xlsx_path = session_dir / "session_data.xlsx"

        try:
            self._generate_xlsx(csv_path, sample_count, weight_history or [])
        except Exception as e:
            print(f"[ExcelRecorder] xlsx generation failed: {e}")
            return None

        return xlsx_path

    def _generate_xlsx(
        self, csv_path: Path, sample_count: int, weight_history: List[tuple]
    ) -> None:
        wb = openpyxl.Workbook(write_only=True)

        # Sheet 1: UR_FT_Data (the full time-series)
        ws1 = wb.create_sheet("UR_FT_Data")
        self._csv_to_xlsx_sheet(csv_path, ws1)

        # Sheet 2: Camera_Meta (from RealSense meta.csv)
        ws2 = wb.create_sheet("Camera_Meta")
        for name in ("meta.csv", "camera_meta.csv"):
            meta_csv = self.session_dir / name
            if meta_csv.exists():
                self._csv_to_xlsx_sheet(meta_csv, ws2)
                break

        # Sheet 3: Session_Info
        ws3 = wb.create_sheet("Session_Info")
        info = [
            ("xlsx_generated_at", datetime.now().isoformat(sep=" ", timespec="seconds")),
            ("session_dir", str(self.session_dir)),
            ("csv_path", str(csv_path)),
            ("sample_count", str(sample_count)),
            ("weight_history", json.dumps(
                [{"time": t, "weight_g": w} for t, w in weight_history],
                ensure_ascii=False,
            )),
        ]
        for key, value in info:
            ws3.append([key, value])

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(str(self.session_dir / "session_data.xlsx"))

    @staticmethod
    def _csv_to_xlsx_sheet(csv_path: Path, ws) -> int:
        count = 0
        with csv_path.open("r", encoding="utf-8") as fh:
            for row in csv.reader(fh):
                ws.append(row)
                count += 1
        return count
